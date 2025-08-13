import csv
from typing import List
from openpyxl import load_workbook

user_profiles = {}
user_results = {}


def read_csv_file(filepath: str):
    """
    读取 CSV 文件并返回列表形式的数据（含表头）
    """
    with open(filepath, newline="", encoding="utf-8") as csvfile:
        reader = csv.reader(csvfile)
        data = [row for row in reader]
    return data


def read_xlsx_file(filepath):
    """
    读取 XLSX 文件并返回列表形式的数据（含表头），默认读取第一个工作表
    """
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data


def load_user_profiles():
    data = read_xlsx_file("../sources/google/user_profiles.xlsx")
    for i in range(len(data)):
        if i == 0:
            continue
        user_profiles[data[i][0]] = data[i][19]


def load_user_results():
    data = read_csv_file("../results/results.csv")
    for i in range(len(data)):
        if i == 0:
            header = data[i]
            continue
        user_results[data[i][1]] = data[i]
    return header


def merge():
    load_user_profiles()
    header = load_user_results()
    header.append("predict")
    # print("--", header)
    # print("++", user_profiles["atsatsj0135798@gmail.com"])
    # print("xx", user_results["atsatsj0135798@gmail.com"])
    rows = []
    for k, v in user_profiles.items():
        if k not in user_results:
            continue
        user_results[k].append(v)
        row = list()
        for v in user_results[k]:
            if v == "<nil>":
                v = "-"
            row.append(v)
        rows.append(row)
    create_csv_file("./merge.csv", header, rows)


def create_csv_file(filepath, header, rows):
    with open(filepath, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(header)
        writer.writerows(rows)

    print(f"CSV 文件已创建：{filepath}")


if __name__ == "__main__":
    merge()
